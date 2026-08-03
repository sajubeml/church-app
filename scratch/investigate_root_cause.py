import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

excel_file = "working Church_Accounting_ok ind updte 21-7-(26-27).xlsm"
print(f"Loading {excel_file}...")

wb_formula = openpyxl.load_workbook(excel_file, data_only=False)
sheet_f = wb_formula['Individual']

wb_val = openpyxl.load_workbook(excel_file, data_only=True)
sheet_v = wb_val['Individual']

print("\n=== ROW 92 FORMULAS VS EVALUATED VALUES IN ORIGINAL EXCEL ===")
for col in range(1, 40):
    col_letter = openpyxl.utils.get_column_letter(col)
    header = sheet_f[f'{col_letter}4'].value
    f_val = sheet_f[f'{col_letter}92'].value
    v_val = sheet_v[f'{col_letter}92'].value
    if f_val is not None or v_val is not None:
        print(f"Cell {col_letter:<2}92 ({str(header):<35}): Formula = {str(f_val):<20} | Evaluated Value = {v_val}")

print("\n=== CASH BOOK SHEET SEARCH FOR SANTHOSH K. A. / REG #150 ===")
cb_f = wb_formula['Cash Book']
cb_v = wb_val['Cash Book']

for r in range(1, cb_f.max_row + 1):
    c_val = str(cb_v.cell(row=r, column=3).value or '')
    d_val = str(cb_v.cell(row=r, column=4).value or '')
    if '150' in c_val or 'santhosh' in d_val.lower():
        row_data = [str(cb_v.cell(row=r, column=c).value or '') for c in range(1, 18)]
        print(f"Cash Book Row {r}: {row_data}")
